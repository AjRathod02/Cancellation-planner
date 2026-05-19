import logging
from io import BytesIO
from typing import Any, Optional

import msal
import openpyxl
import pandas as pd
import requests

from config import get_settings

logger = logging.getLogger(__name__)

_GRAPH_SCOPE = ["https://graph.microsoft.com/.default"]

_DEFAULT_PATHS = {
  "inventory": "Operations/july_2023_email_scraping/2025_wm_inventory_reservations_details.xlsx",
  "confirmed": "Operations/Confirmed Bookings Guest Reservation Tracking.xlsx",
  "credit_mm": "Operations/july_2023_email_scraping/Credit_and_MMaccount.xlsx",
}


def _env_path(name: str, default: str) -> str:
  import os

  value = os.getenv(name, "").strip()
  return value if value else default


def _graph_config() -> dict:
  settings = get_settings()
  return {
    "authority": settings["sp_authority"],
    "client_id": settings["sp_client_id"],
    "secret": settings["sp_client_secret"],
    "scope": _GRAPH_SCOPE,
  }


def get_site_id() -> str:
  return get_settings()["sp_site_id"]


def get_sharepoint_paths() -> dict[str, str]:
  return {
    "inventory": _env_path("SP_INVENTORY_PATH", _DEFAULT_PATHS["inventory"]),
    "confirmed": _env_path("SP_CONFIRMED_PATH", _DEFAULT_PATHS["confirmed"]),
    "credit_mm": _env_path("SP_CREDIT_MM_PATH", _DEFAULT_PATHS["credit_mm"]),
  }


def _acquire_token() -> Optional[str]:
  cfg = _graph_config()
  app = msal.ConfidentialClientApplication(
    client_id=cfg["client_id"],
    authority=cfg["authority"],
    client_credential=cfg["secret"],
  )
  result = app.acquire_token_silent(scopes=cfg["scope"], account=None)
  if not result:
    result = app.acquire_token_for_client(scopes=cfg["scope"])
  if "access_token" not in result:
    logger.error(
      "SharePoint authentication failed: %s - %s",
      result.get("error"),
      result.get("error_description"),
    )
    return None
  return result["access_token"]


def _resolve_sheet(wb: openpyxl.Workbook, sheet_name: int | str | None) -> tuple[int | str, str]:
  sheet = 0 if sheet_name is None else sheet_name
  title = wb.sheetnames[sheet] if isinstance(sheet, int) else sheet
  return sheet, title


def get_sharepoint_file(
  file_path: str,
  sheet_name: int | str | None = 0,
) -> Optional[pd.DataFrame]:
  """Retrieve an Excel file from SharePoint via Microsoft Graph API."""
  try:
    token = _acquire_token()
    if not token:
      return None

    normalized = file_path.lstrip("/")
    site_id = get_site_id()
    endpoint = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{normalized}"

    response = requests.get(
      endpoint,
      headers={"Authorization": f"Bearer {token}"},
      timeout=120,
    )
    response.raise_for_status()
    graph_data = response.json()

    download_url = graph_data.get("@microsoft.graph.downloadUrl")
    if not download_url:
      logger.error("Download URL not found for %s", normalized)
      return None

    file_content = requests.get(download_url, timeout=120).content
    wb = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)
    sheet, title = _resolve_sheet(wb, sheet_name)

    if title in wb.sheetnames:
      ws = wb[title]
      for mrange in list(ws.merged_cells.ranges):
        min_row, min_col, max_row, max_col = (
          mrange.min_row,
          mrange.min_col,
          mrange.max_row,
          mrange.max_col,
        )
        top_left_value = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
          for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).value = top_left_value

      buffer = BytesIO()
      wb.save(buffer)
      buffer.seek(0)
      df = pd.read_excel(buffer, engine="openpyxl", sheet_name=sheet)
    else:
      df = pd.read_excel(BytesIO(file_content), engine="openpyxl", sheet_name=sheet)

    if isinstance(df, dict):
      df = next(iter(df.values()))

    logger.info("Loaded SharePoint file: %s (sheet: %s)", normalized, title)
    return df

  except requests.exceptions.RequestException as e:
    logger.error("Error fetching %s from SharePoint: %s", file_path, e)
    if hasattr(e, "response") and e.response is not None:
      logger.error("Response: %s %s", e.response.status_code, e.response.text[:500])
    return None
  except pd.errors.EmptyDataError:
    logger.error("Excel file is empty or invalid: %s", file_path)
    return None
  except Exception:
    logger.exception("Unexpected error loading %s", file_path)
    return None


def load_sharepoint_inputs() -> dict[str, Any]:
  """Load all Excel inputs required by the cancellation job from SharePoint."""
  paths = get_sharepoint_paths()

  inventory_df = get_sharepoint_file(paths["inventory"], sheet_name=0)
  if inventory_df is None:
    raise RuntimeError(f"Failed to load inventory file: {paths['inventory']}")

  confirmed_df = get_sharepoint_file(paths["confirmed"], sheet_name="upcoming bookings")
  if confirmed_df is None:
    raise RuntimeError(f"Failed to load confirmed bookings file: {paths['confirmed']}")

  credit_df = get_sharepoint_file(paths["credit_mm"], sheet_name="Credits Summary")
  if credit_df is None:
    raise RuntimeError(f"Failed to load credits sheet: {paths['credit_mm']}")

  mm_is_df = get_sharepoint_file(paths["credit_mm"], sheet_name="MM IS resort")
  if mm_is_df is None:
    raise RuntimeError(f"Failed to load MM IS resort sheet: {paths['credit_mm']}")

  return {
    "inventory_df": inventory_df,
    "confirmed_df": confirmed_df,
    "credit_df": credit_df,
    "mm_is_df": mm_is_df,
  }
